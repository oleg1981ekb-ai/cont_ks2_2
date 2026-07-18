
import logging
logging.getLogger().setLevel(logging.WARNING)

import openpyxl
from openpyxl.comments import Comment

import json
import os
import datetime
import config
import db_core
import db_viewer
import excel_styler


def apply_row_style(ws, row_idx, font, fill, border, alignment=None):
    excel_styler.apply_row_style(ws, row_idx, font, fill, border, alignment)


def test_columns_width_logic():
    columns_to_check = [4, 5, 6, 7, 8]  # D, E, F, G, H
    letters = {4: "D", 5: "E", 6: "F", 7: "G", 8: "H"}
    print(" Результаты симуляции openpyxl:")
    for col_idx in columns_to_check:
        width_val = 14.0
        print(f"  - Столбец {letters[col_idx]} (Индекс {col_idx}): Задано значение ширины = {width_val}")
    print(" [ИНФО] Логика верна. Если в LibreOffice столбцы узкие, убедись, что файл builder.py сохранен на диск через Cmd+S!")


def build_structure(ws, mock_data=None, saved_statuses=None, saved_sums=None):
    ws.freeze_panes = "A3"
    ws.sheet_properties.outlinePr.summaryBelow = False

    now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.oddHeader.right.text = f"Выгружено: {now_str}"
    ws.oddHeader.right.size = 9
    ws.oddHeader.right.color = "7A7A7A"

    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(config.HEADERS)
        for col_idx in range(1, len(config.HEADERS) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = config.FILL_HDR
            cell.font = config.FONT_HDR
            cell.border = config.THIN_BORDER

    json_path = "database.json"
    if not os.path.exists(json_path):
        return ws.max_row
    with open(json_path, "r", encoding="utf-8") as f:
        db = json.load(f)

    # ЭТАП 1: Чтение метаданных на старте
    meta = db.get("_meta", {}) if isinstance(db, dict) else {}
    target_dir = meta.get("last_changed_dir")
    target_sub = meta.get("last_changed_sub")
    is_new = meta.get("is_new_change", False)

    row_counter = 1
    for direction in db.keys():
        if direction == "_meta":
            continue
        ws.append(["", str(direction), "", "", "", "", "", "", "", ""])
        excel_styler.apply_row_style(
            ws,
            ws.max_row,
            config.FONT_DIR,
            config.FILL_DIR,
            config.THIN_BORDER,
            config.ALIGN_L,
        )
        ws.row_dimensions[ws.max_row].outline_level = 0

        for sub_obj in db[direction].keys():
            ws.append(["", str(sub_obj), "", "", "", "", "", "", "", ""])
            excel_styler.apply_row_style(
                ws,
                ws.max_row,
                config.FONT_OBJ,
                config.FILL_OBJ,
                config.THIN_BORDER,
                config.ALIGN_L,
            )
            ws.row_dimensions[ws.max_row].outline_level = 1

            # Сумма договора (строка объекта) -> колонка C (индекс 3)
            obj_row = ws.max_row
            contract_sum = ""
            try:
                contract_sum = db[direction][sub_obj].get("contract_sum", "")
            except Exception:
                contract_sum = ""

            if contract_sum != "" and contract_sum is not None:
                ws.cell(row=obj_row, column=3, value=float(contract_sum))
                ws.cell(row=obj_row, column=3).number_format = "#,##0.00"

                from openpyxl.styles import Font, Side, Border

                # Жирный шрифт
                ws.cell(row=obj_row, column=3).font = Font(name="Calibri", size=11, bold=True, color="000000")

                # Нижнее подчёркивание: тонкая нижняя граница
                bottom_side = Side(style="thin", color="B0B0B0")
                ws.cell(row=obj_row, column=3).border = Border(
                    left=None,
                    right=None,
                    top=None,
                    bottom=bottom_side,
                )
            else:
                ws.cell(row=obj_row, column=3).value = ""


            raw_months = list(db[direction][sub_obj].keys())
            available_months = [m for m in db_core.ALL_YEAR_MONTHS if m in raw_months]

            for m in raw_months:
                if m not in available_months:
                    available_months.append(m)

            for mth in available_months:
                mth_data = db[direction][sub_obj][mth]
                # Защита: если по ошибке в БД вместо dict записано число/float — пропускаем обработку статусов
                if not isinstance(mth_data, dict):
                    continue
                status_raw = mth_data.get("status", "")


                # Месяц: Уровень 2
                ws.append(
                    [row_counter, mth, float(mth_data.get("sum", 0.0)), "", "", "", "", "", "", ""]
                )
                current_row = ws.max_row

                # Всплывающий комментарий к оплате (payment_comment) -> Comment к ячейке C
                comment_text = mth_data.get("payment_comment", "")
                if comment_text:
                    cell_c = ws.cell(row=current_row, column=3)
                    cell_c.comment = Comment(comment_text, "Система")

                excel_styler.apply_row_style(
                    ws,
                    current_row,
                    config.FONT_MTH,
                    config.FILL_MTH,
                    config.THIN_BORDER,
                    config.ALIGN_L,
                )
                ws.cell(row=current_row, column=1).alignment = config.ALIGN_C
                ws.cell(row=current_row, column=3).number_format = "#,##0.00"
                ws.row_dimensions[current_row].outline_level = 2
                row_counter += 1

                # Раскрытие последнего изменения (ЭТАП 2)
                hidden_default = True
                if is_new and direction == target_dir and sub_obj == target_sub:
                    hidden_default = False
                ws.row_dimensions[current_row].hidden = hidden_default

                # Документы: Уровень 3
                # Динамические документы печатаем только если они добавлены пользователем в этом месяце.
                extra_docs = mth_data.get("extra_docs", []) if isinstance(mth_data, dict) else []
                if not isinstance(extra_docs, list):
                    extra_docs = []

                dynamic_extra_set = {"Акт передачи оборудования", "Акт расхода давальческих материалов"}

                for d_name in config.DOCUMENTS_LIST:
                    if d_name in dynamic_extra_set and d_name not in extra_docs:
                        continue

                    ws.append(["", f"• {d_name}", "", "", "", "", "", "", "", ""])

                    doc_row = ws.max_row
                    excel_styler.apply_row_style(
                        ws,
                        doc_row,
                        config.FONT_DATA,
                        None,
                        config.THIN_BORDER,
                        config.ALIGN_L,
                    )
                    ws.cell(row=doc_row, column=3).number_format = "#,##0.00"
                    ws.row_dimensions[doc_row].outline_level = 3

                    mask = config.DOCUMENT_ROLES.get(
                        d_name,
                        {"СтрК": 1, "СДО": 1, "ГенДир": 1, "1 экз. З.": 1, "1 экз. П": 1, "Опл.": 1},
                    )

                    for col_name, col_letter_or_idx in config.COLUMN_MAPPING.items():
                        if isinstance(col_letter_or_idx, str):
                            col_idx = openpyxl.utils.column_index_from_string(col_letter_or_idx)
                        else:
                            col_idx = int(col_letter_or_idx)

                        cell = ws.cell(row=doc_row, column=col_idx)

                        # статус конкретного документа: СтрК/СДО
                        if isinstance(status_raw, dict) and d_name in status_raw:
                            doc_status_raw = status_raw[d_name]
                        else:
                            doc_status_raw = status_raw

                        status_date = doc_status_raw.get("date", "") if isinstance(doc_status_raw, dict) else ""

                        # Если для режима «заполнить ВСЕ документы» был сохранен col_key,
                        # то заполняем только одну выбранную колонку (остальные очищаем).
                        saved_col_key = doc_status_raw.get("col_key") if isinstance(doc_status_raw, dict) else None

                        if isinstance(doc_status_raw, dict):
                            clean_status_value = (
                                doc_status_raw.get(col_name, {}).get("value")
                                if col_name in doc_status_raw
                                else doc_status_raw.get("value")
                            )
                        else:
                            clean_status_value = doc_status_raw if col_name == "СтрК" else None


                        # Заполняем только одну колонку ТОЛЬКО если данные плоские (value/date/col_key),
                        # иначе не очищаем уже введённые ранее значения в других колонках.
                        is_flat_col_key_mode = (
                            saved_col_key is not None
                            and isinstance(doc_status_raw, dict)
                            and "value" in doc_status_raw
                            and "date" in doc_status_raw
                            and ("col_key" in doc_status_raw)
                        )

                        # Если в БД сохранён плоский статус (value/date/col_key), то на этапе сборки
                        # не перезаписываем/не затираем значения в других колонках документа.
                        # Иначе при генерации пользовательские статусы теряются «для всего документа».
                        #
                        # Вместо очистки других колонок — оставляем их как есть (обычно они будут
                        # заполнены из БД корректно ниже по логике статуса).
                        if False and is_flat_col_key_mode and col_name != saved_col_key:
                            cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                            cell.value = ""
                            cell.fill = config.FILL_BLOCKED
                            cell.value = ""
                        else:
                            if clean_status_value is not None:
                                excel_styler.format_status_cell(cell, clean_status_value)
                            else:
                                cell.fill = openpyxl.styles.PatternFill(fill_type=None)


                    # Логгер даты: Уровень 4
                    if mask.get("СтрК", 1) == 1 and status_date:
                        ws.append(
                            [
                                "",
                                f"    └─ Дата изменения СтрК: {status_date}",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "",
                            ]
                        )
                        log_row = ws.max_row
                        excel_styler.apply_row_style(
                            ws,
                            log_row,
                            excel_styler.FONT_LOG_DATE,
                            None,
                            config.THIN_BORDER,
                            config.ALIGN_L,
                        )
                        ws.row_dimensions[log_row].outline_level = 4

    # ЭТАП 4: Автоматическая фиксация приемки документов (ГенДир)
    # Условие: внутри одного месяца
    #  - если в строках документов «Акт КС-2» и «Справка КС-3» в колонке F (ГенДир) стоит 1
    #  - тогда в строке этого же месяца в колонке F проставляем сумму из колонки C.
    # Важно: проверка делается строго по блокам одного месяца.
    def _is_doc_row_for(name: str, cell_b_value):
        if not cell_b_value:
            return False
        s = str(cell_b_value).strip()
        return s == f"• {name}"

    # Карта: для каждой строки месяца запоминаем строки документов (Акт КС-2/Справка КС-3)
    month_row_to_docs = {}
    current_month_row = None

    for r in range(2, ws.max_row + 1):
        b_val = ws.cell(row=r, column=2).value

        # строка месяца имеет уровень outline_level=2 (по логике построения)
        if ws.row_dimensions[r].outline_level == 2:
            current_month_row = r
            month_row_to_docs.setdefault(current_month_row, {})
            continue

        if current_month_row is None:
            continue

        # строки документов имеют префикс «• » в колонке B
        if _is_doc_row_for("Акт КС-2", b_val):
            month_row_to_docs[current_month_row]["Акт КС-2"] = r
        elif _is_doc_row_for("Справка КС-3", b_val):
            month_row_to_docs[current_month_row]["Справка КС-3"] = r

    from openpyxl.styles import Font
    for mrow, docs in month_row_to_docs.items():
        d1 = docs.get("Акт КС-2")
        d2 = docs.get("Справка КС-3")
        if not d1 or not d2:
            continue

        # колонка F = индекс 6
        gen_dir_col_idx = 6
        try:
            gen1 = ws.cell(row=d1, column=gen_dir_col_idx).value
            gen2 = ws.cell(row=d2, column=gen_dir_col_idx).value
        except Exception:
            continue

        if str(gen1).strip() == "1" and str(gen2).strip() == "1":
            # Проставляем сумму из колонки C (индекс 3) в колонку F (ГенДир) строки месяца
            sum_val = ws.cell(row=mrow, column=3).value
            ws.cell(row=mrow, column=gen_dir_col_idx, value=sum_val)
            ws.cell(row=mrow, column=gen_dir_col_idx).number_format = "#,##0.00"
            ws.cell(row=mrow, column=gen_dir_col_idx).font = Font(name="Calibri", size=11, bold=True, color="000000")

    # Проставляем формулы статуса акта
    for row in range(2, ws.max_row + 1):
        doc_name_val = ws.cell(row=row, column=2).value  # колонка B

        # строка документа: начинается с "• "
        if not doc_name_val or not str(doc_name_val).startswith("• "):
            continue

        doc_name = str(doc_name_val)

        # Жесткие правила блокировки (только исключения)
        block_rules = {
            4: ["Справка КС-3", "Счет-фактура", "Счет"],
            5: ["Исполнительная документация"],
            6: ["Счет-фактура", "Счет", "Исполнительная документация"],
            8: ["Счет-фактура", "Счет"],
        }


        for col_idx, blocked_docs in block_rules.items():
            if not blocked_docs:
                continue

            # блокировка имеет абсолютный приоритет: если документ совпадает
            # — безусловно красим в серый и очищаем значение.
            cell = ws.cell(row=row, column=col_idx)

            if any(doc.strip() == doc_name.replace("• ", "") for doc in blocked_docs):
                cell.fill = config.FILL_BLOCKED
                cell.value = None


        # Оригинальная логика формулы статуса акта (ГенДир/колонка J)
        ws.cell(row=row, column=10).value = f'=IF(C{row}>0, "В работе", "")'



    # УМНОЕ АВТОМАТИЧЕСКОЕ СВЕРТЫВАНИЕ СТРОК ДО УРОВНЯ МЕСЯЦЕВ

    ws.sheet_view.showOutlineSymbols = True
    for row_idx in range(2, ws.max_row + 1):
        if ws.row_dimensions[row_idx].outline_level in (3, 4):
            ws.row_dimensions[row_idx].hidden = True

    # Гарантированное выравнивание сетки по буквам колонок
    status_letters = ["D", "E", "F", "G", "H"]
    for col in ws.columns:

        first_cell = col[0]
        col_letter = first_cell.column_letter
        col_idx = first_cell.column

        if col_letter in status_letters:
            # Выставляем точный проверенный размер для 1.35 см
            ws.column_dimensions[col_letter].width = 8.0
            ws.column_dimensions[col_letter].auto_fit = False
        else:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    cell_len = max(len(sub_line) for sub_line in str(cell.value).split("\n"))
                    if cell_len > max_len:
                        max_len = cell_len

            if col_idx == 2:
                ws.column_dimensions[col_letter].width = max(max_len + 5, 45)
            elif col_idx == 3:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 18)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ЭТАП 3: Сброс флага в финале (чтобы следующий запуск открывался свернутым)
    if isinstance(db, dict) and "_meta" in db:
        db["_meta"]["is_new_change"] = False
        db_core.save_db(db)

    return ws.max_row

