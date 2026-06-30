from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
import config

def apply_conditional_formatting(ws):
    f_grn = config.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    f_ylw = config.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    f_red = config.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_grn, font_ylw, font_red = Font(name="Calibri", size=11, color="006100"), Font(name="Calibri", size=11, color="9C6500"), Font(name="Calibri", size=11, color="9C0006")
    # Цветовые статусы теперь занимают ячейки D, E, F, G, H
    ws.conditional_formatting.add("D2:H1000", CellIsRule(operator='equal', formula=['1'], fill=f_grn, font=font_grn))
    ws.conditional_formatting.add("D2:H1000", CellIsRule(operator='equal', formula=['2'], fill=f_ylw, font=font_ylw))
    ws.conditional_formatting.add("D2:H1000", CellIsRule(operator='equal', formula=['3'], fill=f_red, font=font_red))

def append_specifications(ws, current_row):
    current_row += 2
    ws.cell(row=current_row, column=2, value="СПЕЦИФИКАЦИЯ ОБОЗНАЧЕНИЙ:").font = Font(name="Calibri", size=11, bold=True)
    specs = [(3, "Еще не началось согласование (Красный цвет)"), (2, "В работе / На согласовании (Желтый цвет)"), (1, "Документы согласованы (Зеленый цвет)")]
    for score, text in specs:
        current_row += 1
        ws.cell(row=current_row, column=2, value=text).font = config.FONT_DATA
        sc = ws.cell(row=current_row, column=4, value=score)
        sc.font, sc.alignment, sc.border = config.FONT_DATA, config.ALIGN_C, config.THIN_BORDER
    return current_row

def set_column_widths(ws):
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 8, 55, 16
    ws.column_dimensions['D'].width = ws.column_dimensions['E'].width = ws.column_dimensions['F'].width = ws.column_dimensions['G'].width = ws.column_dimensions['H'].width = 7
    # Столбец I (Опл.) делаем компактным (ширина 10)
    ws.column_dimensions['I'].width = 10
    # Столбец J (Текущий статус акта) делаем широким (ширина 40)
    ws.column_dimensions['J'].width = 40
