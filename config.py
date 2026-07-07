import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DATA_ROWS = [
    ('01_АНАПА (Таманская)', '01_271_КН', 'Февраль', 450000, 3),
    ('01_АНАПА (Таманская)', '01_271_КН', 'Март', 0, None),
    ('01_АНАПА (Таманская)', '02_1505_КН.', 'Февраль', 450000, 3),
    ('01_АНАПА (Таманская)', '02_1505_КН.', 'Март', 0, None),
    ('01_АНАПА (Таманская)', '03_538_КН', 'Февраль', 0, None),
    ('01_АНАПА (Таманская)', '03_538_КН', 'Март', 0, None),
    ('01_АНАПА (Таманская)', '04_537_КН', 'Февраль', 0, None),
    ('01_АНАПА (Таманская)', '04_537_КН', 'Март', 0, None),
    ('02_ИНДУСТРИАЛЬНАЯ Краснодар', 'Основной участок', 'Февраль', 850000, 2),
    ('02_ИНДУСТРИАЛЬНАЯ Краснодар', 'Основной участок', 'Март', 0, None),
]

MONTHS_LIST = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль']
DOCUMENTS_LIST = ['Акт КС-2', 'Справка КС-3', 'Счет-фактура', 'Счет', 'Исполнительная документация']

FOLDER_NAME = ""
FILE_NAME = "Трекер_Акт_Выполнения.xlsx"
FULL_PATH = os.path.join(FOLDER_NAME, FILE_NAME)

FONT_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_DIR = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
FONT_OBJ = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_MTH = Font(name="Calibri", size=11, bold=True, italic=True, color="000000")
FONT_DATA = Font(name="Calibri", size=11)

FILL_HDR = FILL_DIR = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_OBJ = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_MTH = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")

HEADERS = [
    "№ п/п",
    "Наименование объекта / Месяц / Документ",
    "Сумма (руб.)",
    "СтрК",
    "СДО",
    "ГенДир",
    "1 экз. З.",
    "1 экз. П",
    "Опл.",
    "Текущий статус акта",
]

# === МАСКИ ПОВЕДЕНИЯ (ТРАФАРЕТЫ ДЛЯ ДОКУМЕНТОВ) ===
# 1 - колонка активна (белая), 0 - заблокирована (серая заливка)
ROLE_KS2      = {"СтрК": 1, "СДО": 1, "ГенДир": 1, "1 экз. З.": 1, "1 экз. П": 1, "Опл.": 0}
ROLE_KS3      = {"СтрК": 0, "СДО": 1, "ГенДир": 1, "1 экз. З.": 1, "1 экз. П": 1, "Опл.": 0}
ROLE_INVOICE  = {"СтрК": 0, "СДО": 1, "ГенДир": 0, "1 экз. З.": 1, "1 экз. П": 1, "Опл.": 0}
ROLE_EXEC_DOC = {"СтрК": 1, "СДО": 0, "ГенДир": 0, "1 экз. З.": 1, "1 экз. П": 1, "Опл.": 0}


DOCUMENT_ROLES = {
    "Акт КС-2": ROLE_KS2,
    "Справка КС-3": ROLE_KS3,
    "Счет-фактура": ROLE_INVOICE,
    "Счет": ROLE_INVOICE,
    "Исполнительная документация": ROLE_EXEC_DOC,
}

# Пастельно-серый цвет для блокировки неиспользуемых ячеек
FILL_BLOCKED = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

# === ОТЛИЧИЯ СДО ДЛЯ СМЕЖНЫХ КОЛОНОК ===
# СтрК: D, СДО: E, ГенДир: F
COLUMN_MAPPING = {
    "СтрК": "D",
    "СДО": "E",
    "ГенДир": "F",
}

# === МАСКИ РОЛЕЙ СТАТУСОВ ===
# Маска для СтрК уже описана в таблице ROLE_KS2/ROLE_KS3/ROLE_INVOICE/ROLE_EXEC_DOC.
# Для новой колонки "ГенДир" используем маску ролей по аналогии с СДО.
ROLE_GEN_DIR = {
    "Акт КС-2": 1,
    "Справка КС-3": 1,
    "Счет-фактура": 0,
    "Счет": 1,
    "Исполнительная документация": 0,
}

# Общий реестр масок по типу статуса
ROLES = {
    "СДО": ROLE_KS2,  # СДО уже управляется базовыми масками в ROLE_KS2/ROLE_KS3 и документ-ролями
    "ГенДир": ROLE_GEN_DIR,
}

VERSION = "v1.3.1"



# Дополнительные цвета для автоматической покраски статусов СтрК
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# === ПОДДЕРЖКА СДО (аналогично СтрК) ===
# Если захотите отличающуюся палитру для СДО — меняйте сюда.
FILL_SDO_GREEN = FILL_GREEN
FILL_SDO_YELLOW = FILL_YELLOW
FILL_SDO_RED = FILL_RED

# Цвета для статусов СДО (аналогично СтрК)
SDO_STATUS_COLORS = {
    "green": FILL_SDO_GREEN,
    "yellow": FILL_SDO_YELLOW,
    "red": FILL_SDO_RED,
}

