import os
import openpyxl
from config import MONTHS_LIST

def load_historical_data(file_path):
    saved_statuses, saved_sums = {}, {}
    # Если файла нет — сразу выходим без ошибок и создаем с нуля
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        return saved_statuses, saved_sums
    try:
        old_wb = openpyxl.load_workbook(file_path, data_only=True)
        if "Реестр Актов" not in old_wb.sheetnames:
            old_wb.close()
            return saved_statuses, saved_sums
        old_ws = old_wb["Реестр Актов"]
        c_dir, c_sub, c_mth = "", "", ""
        for row in range(2, old_ws.max_row + 1):
            val_b = str(old_ws.cell(row=row, column=2).value or "").strip()
            if not val_b: continue
            if "01_АНАПА" in val_b: c_dir = "01_АНАПА (Таманская)"
            elif "02_ИНДУСТРИАЛЬНАЯ" in val_b: c_dir = "02_ИНДУСТРИАЛЬНАЯ Краснодар"
            elif val_b.startswith("Подобъект "): c_sub = val_b.replace("Подобъект ", "").strip()
            elif val_b in MONTHS_LIST: c_mth = val_b
            elif val_b.startswith("• "):
                d_name = val_b.replace("• ", "").strip()
                st_d = old_ws.cell(row=row, column=4).value
                st_e = old_ws.cell(row=row, column=5).value
                st_f = old_ws.cell(row=row, column=6).value
                st_g = old_ws.cell(row=row, column=7).value
                # Сохраняем все 4 статуса контроля
                if any(st in (1, 2, 3) for st in (st_d, st_e, st_f, st_g)):
                    saved_statuses[(c_dir, c_sub, c_mth, d_name)] = (st_d, st_e, st_f, st_g)
                if d_name == "Акт КС-2":
                    s_val = old_ws.cell(row=row, column=3).value
                    if s_val is not None: saved_sums[(c_dir, c_sub, c_mth)] = s_val
        old_wb.close()
    except Exception: 
        pass
    return saved_statuses, saved_sums

def group_linear_data(data_rows):
    mock_data = {}
    for direction, sub_obj, month, val, status in data_rows:
        if direction not in mock_data: mock_data[direction] = {}
        if sub_obj not in mock_data[direction]: mock_data[direction][sub_obj] = {}
        mock_data[direction][sub_obj][month] = [val, status]
    return mock_data
